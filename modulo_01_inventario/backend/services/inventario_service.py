from datetime import datetime
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from backend.models.models import Alerta, Movimiento, NivelAlertaEnum, Producto, TipoMovimientoEnum


def clasificar_producto(producto: Producto) -> str:
    if producto.stock == 0:
        return "agotado"
    if producto.stock < producto.stock_minimo:
        return "bajo"
    return "normal"


def clasificar_todos(db: Session) -> List[Dict]:
    orden = {"agotado": 0, "bajo": 1, "normal": 2}
    resultado = []

    for producto in db.query(Producto).all():
        estado = clasificar_producto(producto)
        resultado.append(
            {
                "id": producto.id,
                "nombre": producto.nombre,
                "categoria": producto.categoria.value,
                "stock": producto.stock,
                "stock_minimo": producto.stock_minimo,
                "precio": producto.precio,
                "valor_total": round(producto.stock * producto.precio, 2),
                "estado": estado,
                "unidad": producto.unidad,
            }
        )

    resultado.sort(key=lambda item: orden[item["estado"]])
    return resultado


def detectar_alertas_stock(db: Session) -> List[Dict]:
    alertas_generadas = []

    for producto in db.query(Producto).all():
        estado = clasificar_producto(producto)
        if estado not in ("agotado", "bajo"):
            continue

        alerta_existente = (
            db.query(Alerta)
            .filter(Alerta.producto_id == producto.id, Alerta.resuelta == 0)
            .first()
        )
        if alerta_existente:
            continue

        if estado == "agotado":
            nivel = NivelAlertaEnum.critico
            mensaje = f"CRITICO: '{producto.nombre}' esta agotado. Reabastecimiento urgente."
        else:
            faltantes = producto.stock_minimo - producto.stock
            nivel = NivelAlertaEnum.bajo
            mensaje = (
                f"STOCK BAJO: '{producto.nombre}' tiene {producto.stock} unidades "
                f"(minimo: {producto.stock_minimo}). Faltan {faltantes} unidades."
            )

        db.add(
            Alerta(
                producto_id=producto.id,
                nivel=nivel,
                mensaje=mensaje,
            )
        )
        alertas_generadas.append(
            {
                "producto": producto.nombre,
                "estado": estado,
                "nivel": nivel.value,
                "mensaje": mensaje,
                "stock_actual": producto.stock,
                "stock_minimo": producto.stock_minimo,
            }
        )

    db.commit()
    return alertas_generadas


def detectar_poca_rotacion(db: Session, dias_limite: int = 30) -> List[Dict]:
    poca_rotacion = []

    for producto in db.query(Producto).all():
        ultimo_movimiento = (
            db.query(Movimiento)
            .filter(Movimiento.producto_id == producto.id)
            .order_by(Movimiento.creado_en.desc())
            .first()
        )

        if ultimo_movimiento is None:
            dias_sin_movimiento = None
            condicion = "sin_historial"
        else:
            dias_sin_movimiento = (datetime.utcnow() - ultimo_movimiento.creado_en.replace(tzinfo=None)).days
            condicion = "estancado" if dias_sin_movimiento >= dias_limite else None

        if condicion is None:
            continue

        poca_rotacion.append(
            {
                "id": producto.id,
                "nombre": producto.nombre,
                "categoria": producto.categoria.value,
                "stock": producto.stock,
                "dias_sin_movimiento": dias_sin_movimiento,
                "ultimo_movimiento": ultimo_movimiento.creado_en.strftime("%Y-%m-%d") if ultimo_movimiento else "Nunca",
                "condicion": condicion,
                "valor_estancado": round(producto.stock * producto.precio, 2),
            }
        )

        alerta_existente = (
            db.query(Alerta)
            .filter(
                Alerta.producto_id == producto.id,
                Alerta.nivel == NivelAlertaEnum.sin_rotacion,
                Alerta.resuelta == 0,
            )
            .first()
        )
        if not alerta_existente:
            periodo = "sin historial" if dias_sin_movimiento is None else f"{dias_sin_movimiento} dias"
            db.add(
                Alerta(
                    producto_id=producto.id,
                    nivel=NivelAlertaEnum.sin_rotacion,
                    mensaje=f"SIN ROTACION: '{producto.nombre}' lleva {periodo} sin movimiento. Stock inmovilizado: {producto.stock} unidades.",
                )
            )

    db.commit()
    poca_rotacion.sort(
        key=lambda item: item["dias_sin_movimiento"] if item["dias_sin_movimiento"] is not None else 9999,
        reverse=True,
    )
    return poca_rotacion


def registrar_movimiento(
    db: Session,
    producto_id: int,
    tipo: str,
    cantidad: int,
    nota: str = "",
) -> Tuple[bool, str, Dict]:
    producto = db.query(Producto).filter(Producto.id == producto_id).first()
    if not producto:
        return False, f"Producto con ID {producto_id} no encontrado.", {}

    if cantidad <= 0:
        return False, "La cantidad debe ser mayor a 0.", {}

    if tipo == "salida" and cantidad > producto.stock:
        return False, (
            f"Stock insuficiente. Disponible: {producto.stock} unidades, solicitado: {cantidad} unidades."
        ), {}

    stock_anterior = producto.stock

    if tipo == "entrada":
        stock_nuevo = stock_anterior + cantidad
    elif tipo == "salida":
        stock_nuevo = stock_anterior - cantidad
    elif tipo == "ajuste":
        stock_nuevo = cantidad
    else:
        return False, f"Tipo de movimiento '{tipo}' no válido.", {}

    producto.stock = stock_nuevo
    db.add(
        Movimiento(
            producto_id=producto_id,
            tipo=TipoMovimientoEnum(tipo),
            cantidad=cantidad,
            stock_anterior=stock_anterior,
            stock_nuevo=stock_nuevo,
            nota=nota,
        )
    )
    db.flush()
    detectar_alertas_stock(db)
    db.commit()

    return True, f"Movimiento registrado correctamente. Stock: {stock_anterior} -> {stock_nuevo}", {
        "producto": producto.nombre,
        "tipo": tipo,
        "cantidad": cantidad,
        "stock_anterior": stock_anterior,
        "stock_nuevo": stock_nuevo,
        "fecha": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    }


def exportar_a_excel(db: Session, ruta_salida: str = "reporte_sipu.xlsx") -> str:
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    fill_red = PatternFill("solid", fgColor="D62828")
    fill_dark = PatternFill("solid", fgColor="404040")
    fill_low = PatternFill("solid", fgColor="FFF3CD")
    fill_critical = PatternFill("solid", fgColor="FFCCCC")
    fill_ok = PatternFill("solid", fgColor="D4EDDA")
    center = Alignment(horizontal="center")

    ws = wb.active
    ws.title = "Inventario"
    headers = ["ID", "Nombre", "Categoria", "Stock", "Stock Min.", "Precio", "Valor Total", "Estado", "Unidad"]
    for column, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column, value=title)
        cell.font = header_font
        cell.fill = fill_red
        cell.alignment = center

    for row, producto in enumerate(db.query(Producto).all(), 2):
        estado = clasificar_producto(producto)
        values = [
            producto.id,
            producto.nombre,
            producto.categoria.value,
            producto.stock,
            producto.stock_minimo,
            producto.precio,
            round(producto.stock * producto.precio, 2),
            estado,
            producto.unidad,
        ]
        color = fill_critical if estado == "agotado" else fill_low if estado == "bajo" else fill_ok
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.fill = color
            if column in (4, 5, 6, 7):
                cell.alignment = center

    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=10)
        ws.column_dimensions[column[0].column_letter].width = max_length + 4

    ws_mov = wb.create_sheet("Movimientos")
    headers_mov = ["ID", "Producto", "Tipo", "Cantidad", "Stock Anterior", "Stock Nuevo", "Nota", "Fecha"]
    for column, title in enumerate(headers_mov, 1):
        cell = ws_mov.cell(row=1, column=column, value=title)
        cell.font = header_font
        cell.fill = fill_dark
        cell.alignment = center

    for row, movimiento in enumerate(db.query(Movimiento).order_by(Movimiento.creado_en.desc()).limit(500).all(), 2):
        ws_mov.cell(row=row, column=1, value=movimiento.id)
        ws_mov.cell(row=row, column=2, value=movimiento.producto.nombre if movimiento.producto else "-")
        ws_mov.cell(row=row, column=3, value=movimiento.tipo.value)
        ws_mov.cell(row=row, column=4, value=movimiento.cantidad)
        ws_mov.cell(row=row, column=5, value=movimiento.stock_anterior)
        ws_mov.cell(row=row, column=6, value=movimiento.stock_nuevo)
        ws_mov.cell(row=row, column=7, value=movimiento.nota or "")
        ws_mov.cell(row=row, column=8, value=str(movimiento.creado_en))

    for column in ws_mov.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=10)
        ws_mov.column_dimensions[column[0].column_letter].width = max_length + 4

    ws_alertas = wb.create_sheet("Alertas")
    headers_alertas = ["ID", "Producto", "Nivel", "Mensaje", "Resuelta", "Fecha"]
    for column, title in enumerate(headers_alertas, 1):
        cell = ws_alertas.cell(row=1, column=column, value=title)
        cell.font = header_font
        cell.fill = fill_red
        cell.alignment = center

    for row, alerta in enumerate(db.query(Alerta).order_by(Alerta.creado_en.desc()).all(), 2):
        ws_alertas.cell(row=row, column=1, value=alerta.id)
        ws_alertas.cell(row=row, column=2, value=alerta.producto.nombre if alerta.producto else "-")
        ws_alertas.cell(row=row, column=3, value=alerta.nivel.value)
        ws_alertas.cell(row=row, column=4, value=alerta.mensaje)
        ws_alertas.cell(row=row, column=5, value="Si" if alerta.resuelta else "No")
        ws_alertas.cell(row=row, column=6, value=str(alerta.creado_en))

    for column in ws_alertas.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=10)
        ws_alertas.column_dimensions[column[0].column_letter].width = max_length + 4

    wb.save(ruta_salida)
    return ruta_salida